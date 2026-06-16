#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass, replace
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl


TIMEZONE_ID = "Asia/Taipei"
STORE_NAME = "吉康藥局"
REST_COLOR = "#D50000"
ALL_PEOPLE_NAME = "全部人"
ALL_PEOPLE_FILENAME = "all.ics"
PERSON_SLUGS = {
    "宏志": "hongzhi",
    "佳惠": "jiahui",
    "盈萱": "yingxuan",
    "重琦": "zhongqi",
    "鈺茜": "yuxi",
    "士昂": "shiang",
    "恩婕": "enjie",
}


@dataclass(frozen=True)
class ShiftPart:
    summary: str
    start: time | None = None
    end: time | None = None
    all_day: bool = False
    color: str | None = None
    transparent: bool = False


@dataclass(frozen=True)
class ShiftEvent:
    person: str
    work_date: date
    code: str
    summary: str
    source: str
    start: time | None = None
    end: time | None = None
    all_day: bool = False
    color: str | None = None
    transparent: bool = False
    part_index: int = 0


def t(hour: int, minute: int = 0) -> time:
    return time(hour, minute)


SHIFT_DEFINITIONS: dict[str, list[ShiftPart]] = {
    "M": [ShiftPart("早班", t(8), t(17))],
    "M8": [ShiftPart("早班8-13", t(8), t(13))],
    "M9": [ShiftPart("早班8.5-17.5", t(8, 30), t(17, 30))],
    "M10": [ShiftPart("早班10-18", t(10), t(18))],
    "M15": [ShiftPart("早班8-15", t(8), t(15))],
    "M16": [ShiftPart("早班8-16", t(8), t(16))],
    "E": [ShiftPart("晚班", t(17), t(22))],
    "E13": [ShiftPart("晚班13-22", t(13), t(22))],
    "E14": [ShiftPart("晚班14-22", t(14), t(22))],
    "E15": [ShiftPart("晚班15-22", t(15), t(22))],
    "E16": [ShiftPart("晚班16-22", t(16), t(22))],
    "A": [ShiftPart("整天", t(12), t(22))],
    "S": [
        ShiftPart("二頭班 8-13，17-21", t(8), t(13)),
        ShiftPart("二頭班 8-13，17-21", t(17), t(21)),
    ],
    "O": [ShiftPart("休息", all_day=True, color=REST_COLOR, transparent=True)],
    # Extra codes already present in the workbook.
    "早": [ShiftPart("早：早班 8.5-13", t(8, 30), t(13))],
    "午": [ShiftPart("午：午班 13-17", t(13), t(17))],
    "晚": [ShiftPart("晚：晚班 17-21", t(17), t(21))],
}

SKIP_CODES = {"", "X", "Ｘ", "-", "—", "–"}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_shift_code(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return text.upper() if re.fullmatch(r"[A-Za-z]\d*", text) else text


def as_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and int(value) == value:
        return int(value)
    text = normalize_text(value)
    if re.fullmatch(r"\d+", text):
        return int(text)
    return None


def roc_to_ad_year(value: Any) -> int | None:
    year = as_int(value)
    if year is None:
        return None
    return year + 1911 if year < 1911 else year


def parse_hour(value: str) -> time:
    number = float(value)
    hour = int(number)
    minute = round((number - hour) * 60)
    if minute == 60:
        hour += 1
        minute = 0
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError(f"Invalid hour: {value}")
    return time(hour, minute)


def format_hour(value: str) -> str:
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return str(number).rstrip("0").rstrip(".")


def numeric_shift_parts(code: str) -> list[ShiftPart] | None:
    match = re.fullmatch(
        r"(\d+(?:\.\d+)?)\s*(?:-|~|～|－|到|\s)\s*(\d+(?:\.\d+)?)", code
    )
    if not match:
        return None
    start_text, end_text = match.groups()
    label = f"{format_hour(start_text)}-{format_hour(end_text)}"
    return [ShiftPart(label, parse_hour(start_text), parse_hour(end_text))]


def resolve_shift(code: str) -> list[ShiftPart] | None:
    if code in SHIFT_DEFINITIONS:
        return SHIFT_DEFINITIONS[code]
    return numeric_shift_parts(code)


def excel_cell(row: int, col: int) -> str:
    return f"{openpyxl.utils.get_column_letter(col)}{row}"


def iter_schedule_events(workbook_path: Path) -> tuple[list[ShiftEvent], list[str]]:
    warnings: list[str] = []
    events: list[ShiftEvent] = []
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    for worksheet in workbook.worksheets:
        year = roc_to_ad_year(worksheet["V1"].value)
        month = as_int(worksheet["Z1"].value)
        if year is None or month is None:
            warnings.append(f"{workbook_path.name}/{worksheet.title}: 找不到 V1 年份或 Z1 月份，已略過。")
            continue

        date_columns: list[tuple[int, date]] = []
        for col in range(2, worksheet.max_column + 1):
            day = as_int(worksheet.cell(2, col).value)
            if day is None:
                continue
            try:
                work_date = date(year, month, day)
            except ValueError:
                continue
            date_columns.append((col, work_date))

        if not date_columns:
            warnings.append(f"{workbook_path.name}/{worksheet.title}: 找不到日期欄，已略過。")
            continue

        header_row = None
        for row in range(1, min(worksheet.max_row, 12) + 1):
            if normalize_text(worksheet.cell(row, 1).value) == "人員":
                header_row = row
                break
        if header_row is None:
            warnings.append(f"{workbook_path.name}/{worksheet.title}: 找不到人員列，已略過。")
            continue

        for row in range(header_row + 1, worksheet.max_row + 1):
            person = normalize_text(worksheet.cell(row, 1).value)
            if not person:
                break
            for col, work_date in date_columns:
                code = normalize_shift_code(worksheet.cell(row, col).value)
                if code in SKIP_CODES:
                    continue
                parts = resolve_shift(code)
                if parts is None:
                    source = f"{workbook_path.name}/{worksheet.title}!{excel_cell(row, col)}"
                    warnings.append(f"{source}: 未定義班別 `{code}`，已略過。")
                    continue
                for index, part in enumerate(parts, start=1):
                    source = f"{workbook_path.name}/{worksheet.title}!{excel_cell(row, col)}"
                    events.append(
                        ShiftEvent(
                            person=person,
                            work_date=work_date,
                            code=code,
                            summary=part.summary,
                            start=part.start,
                            end=part.end,
                            all_day=part.all_day,
                            color=part.color,
                            transparent=part.transparent,
                            source=source,
                            part_index=index,
                        )
                    )

    return events, warnings


def ical_escape(value: str) -> str:
    return (
        value.replace("\\", "\\\\")
        .replace("\n", "\\n")
        .replace(";", "\\;")
        .replace(",", "\\,")
    )


def fold_ical_line(line: str) -> str:
    output: list[str] = []
    current = ""
    current_bytes = 0

    for char in line:
        char_bytes = len(char.encode("utf-8"))
        limit = 75
        if current and current_bytes + char_bytes > limit:
            output.append(current)
            current = " " + char
            current_bytes = 1 + char_bytes
        else:
            current += char
            current_bytes += char_bytes

    output.append(current)
    return "\r\n".join(output)


def write_text_file(path: Path, content: str) -> None:
    with path.open("w", encoding="utf-8", newline="") as output:
        output.write(content)


def format_datetime(value: datetime) -> str:
    return value.strftime("%Y%m%dT%H%M%S")


def format_date(value: date) -> str:
    return value.strftime("%Y%m%d")


def event_uid(event: ShiftEvent) -> str:
    raw = f"{event.person}|{event.work_date.isoformat()}|{event.code}|{event.part_index}"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()
    return f"{digest}@jikang-shift-calendar"


def add_line(lines: list[str], name: str, value: str) -> None:
    lines.append(fold_ical_line(f"{name}:{value}"))


def render_calendar(person: str, events: list[ShiftEvent], generated_at: datetime) -> str:
    stamp = generated_at.astimezone(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Jikang Pharmacy//Excel Shift ICS//ZH-TW",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-PUBLISHED-TTL:PT1H",
        "REFRESH-INTERVAL;VALUE=DURATION:PT1H",
    ]
    add_line(lines, "X-WR-CALNAME", ical_escape(f"{STORE_NAME}班表 - {person}"))
    add_line(lines, "X-WR-TIMEZONE", TIMEZONE_ID)
    lines.extend(
        [
            "BEGIN:VTIMEZONE",
            f"TZID:{TIMEZONE_ID}",
            f"X-LIC-LOCATION:{TIMEZONE_ID}",
            "BEGIN:STANDARD",
            "TZOFFSETFROM:+0800",
            "TZOFFSETTO:+0800",
            "TZNAME:CST",
            "DTSTART:19700101T000000",
            "END:STANDARD",
            "END:VTIMEZONE",
        ]
    )

    for event in sorted(events, key=lambda item: (item.work_date, item.start or time.min, item.summary)):
        lines.append("BEGIN:VEVENT")
        add_line(lines, "UID", event_uid(event))
        add_line(lines, "DTSTAMP", stamp)
        add_line(lines, "LAST-MODIFIED", stamp)
        add_line(lines, "SUMMARY", ical_escape(event.summary))
        add_line(
            lines,
            "DESCRIPTION",
            ical_escape(f"人員：{event.person}\n班別：{event.code}\n來源：{event.source}"),
        )
        add_line(lines, "LOCATION", ical_escape(STORE_NAME))
        add_line(lines, "STATUS", "CONFIRMED")
        add_line(lines, "TRANSP", "TRANSPARENT" if event.transparent else "OPAQUE")
        add_line(lines, "CATEGORIES", ical_escape("休息" if event.code == "O" else "班表"))
        if event.color:
            add_line(lines, "COLOR", event.color)
        if event.all_day:
            add_line(lines, "DTSTART;VALUE=DATE", format_date(event.work_date))
            add_line(lines, "DTEND;VALUE=DATE", format_date(event.work_date + timedelta(days=1)))
        else:
            if event.start is None or event.end is None:
                raise ValueError(f"Timed event is missing start/end: {event}")
            start_dt = datetime.combine(event.work_date, event.start)
            end_dt = datetime.combine(event.work_date, event.end)
            if end_dt <= start_dt:
                end_dt += timedelta(days=1)
            add_line(lines, f"DTSTART;TZID={TIMEZONE_ID}", format_datetime(start_dt))
            add_line(lines, f"DTEND;TZID={TIMEZONE_ID}", format_datetime(end_dt))
        lines.append("END:VEVENT")

    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


def with_person_in_summary(events: list[ShiftEvent]) -> list[ShiftEvent]:
    return [
        replace(event, summary=f"{event.person}｜{event.summary}", part_index=event.part_index + 1000)
        for event in events
    ]


def slugify_person(name: str, used: set[str]) -> str:
    base = PERSON_SLUGS.get(name)
    if base is None:
        base = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
        base = re.sub(r"[^a-zA-Z0-9]+", "-", base).strip("-").lower()
    if not base:
        base = "person-" + hashlib.sha1(name.encode("utf-8")).hexdigest()[:8]

    slug = base
    counter = 2
    while slug in used:
        slug = f"{base}-{counter}"
        counter += 1
    used.add(slug)
    return slug


def find_workbooks(input_dir: Path, explicit_files: list[str]) -> list[Path]:
    if explicit_files:
        return [Path(item).expanduser().resolve() for item in explicit_files]
    return sorted(
        path
        for path in input_dir.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    )


def render_index(entries: list[dict[str, Any]], generated_at: datetime, base_url: str) -> str:
    rows = []
    for entry in entries:
        url = entry["url"] or entry["path"]
        rows.append(
            "<tr>"
            f"<td>{html.escape(entry['person'])}</td>"
            f"<td>{entry['events']}</td>"
            f"<td><a href=\"{html.escape(url)}\">{html.escape(url)}</a></td>"
            "</tr>"
        )
    base_note = (
        f"<p>訂閱網址基底：<code>{html.escape(base_url)}</code></p>"
        if base_url
        else "<p>部署到 HTTPS 靜態網站後，將本頁連結換成網站上的完整網址即可訂閱。</p>"
    )
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(STORE_NAME)}班表 ICS</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 32px; line-height: 1.5; }}
    table {{ border-collapse: collapse; width: 100%; max-width: 980px; }}
    th, td {{ border: 1px solid #d0d7de; padding: 8px 10px; text-align: left; }}
    th {{ background: #f6f8fa; }}
    code {{ background: #f6f8fa; padding: 2px 4px; border-radius: 4px; }}
  </style>
</head>
<body>
  <h1>{html.escape(STORE_NAME)}班表 ICS</h1>
  <p>產生時間：{html.escape(generated_at.strftime("%Y-%m-%d %H:%M:%S UTC"))}</p>
  {base_note}
  <table>
    <thead>
      <tr><th>人員</th><th>活動數</th><th>ICS 訂閱網址</th></tr>
    </thead>
    <tbody>
      {''.join(rows)}
    </tbody>
  </table>
</body>
</html>
"""


def write_outputs(
    output_dir: Path,
    events: list[ShiftEvent],
    warnings: list[str],
    base_url: str,
    generated_at: datetime,
) -> dict[str, Any]:
    ics_dir = output_dir / "ics"
    ics_dir.mkdir(parents=True, exist_ok=True)
    for old_file in ics_dir.glob("*.ics"):
        old_file.unlink()

    by_person: dict[str, list[ShiftEvent]] = {}
    for event in events:
        by_person.setdefault(event.person, []).append(event)

    used_slugs: set[str] = set()
    entries: list[dict[str, Any]] = []
    for person in sorted(by_person):
        person_events = by_person[person]
        slug = slugify_person(person, used_slugs)
        filename = f"{slug}.ics"
        calendar_text = render_calendar(person, person_events, generated_at)
        write_text_file(ics_dir / filename, calendar_text)
        path = f"ics/{filename}"
        entries.append(
            {
                "person": person,
                "events": len(person_events),
                "file": filename,
                "path": path,
                "url": f"{base_url.rstrip('/')}/{path}" if base_url else "",
            }
        )

    if events:
        all_people_path = f"ics/{ALL_PEOPLE_FILENAME}"
        write_text_file(
            ics_dir / ALL_PEOPLE_FILENAME,
            render_calendar(ALL_PEOPLE_NAME, with_person_in_summary(events), generated_at),
        )
        entries.insert(
            0,
            {
                "person": ALL_PEOPLE_NAME,
                "events": len(events),
                "file": ALL_PEOPLE_FILENAME,
                "path": all_people_path,
                "url": f"{base_url.rstrip('/')}/{all_people_path}" if base_url else "",
            },
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    report = {
        "generated_at": generated_at.isoformat(),
        "timezone": TIMEZONE_ID,
        "people": entries,
        "warnings": warnings,
    }
    (output_dir / "calendars.json").write_text(
        json.dumps(entries, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (output_dir / "sync-report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (output_dir / "index.html").write_text(
        render_index(entries, generated_at, base_url), encoding="utf-8"
    )
    return report


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Excel 班表轉每人 ICS 訂閱檔。")
    parser.add_argument("--input-dir", default=str(script_dir), help="放置班表 .xlsx 的資料夾。")
    parser.add_argument("--output-dir", default=str(script_dir / "public"), help="輸出靜態網站資料夾。")
    parser.add_argument("--base-url", default=os.environ.get("ICS_BASE_URL", ""), help="部署後的 HTTPS 網址基底。")
    parser.add_argument("workbooks", nargs="*", help="指定要讀取的 .xlsx；省略時讀取 input-dir 內全部 .xlsx。")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    workbooks = find_workbooks(input_dir, args.workbooks)
    if not workbooks:
        print(f"找不到 .xlsx 班表：{input_dir}", file=sys.stderr)
        return 1

    all_events: list[ShiftEvent] = []
    all_warnings: list[str] = []
    for workbook in workbooks:
        events, warnings = iter_schedule_events(workbook)
        all_events.extend(events)
        all_warnings.extend(warnings)

    generated_at = datetime.now(timezone.utc)
    report = write_outputs(output_dir, all_events, all_warnings, args.base_url.strip(), generated_at)

    print(f"已產生 {len(report['people'])} 人的 ICS。")
    for person in report["people"]:
        target = person["url"] or person["path"]
        print(f"- {person['person']}: {target} ({person['events']} events)")
    if all_warnings:
        print(f"警告：{len(all_warnings)} 筆，詳見 {output_dir / 'sync-report.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
